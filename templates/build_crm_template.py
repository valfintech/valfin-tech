from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = "/Users/kejsicenuka/Desktop/Valfin Tech/templates/Roofing_CRM_Google_Sheets_TEMPLATE.xlsx"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="2F5496")
NOTE_FONT = Font(name="Arial", italic=True, color="808080", size=9)
BODY_FONT = Font(name="Arial", size=10)

wb = Workbook()
wb.remove(wb.active)

def make_sheet(name, headers, rows, col_widths=None, note=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append(r)
    for r in range(2, 2 + len(rows)):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
    widths = col_widths or [18] * len(headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    if note:
        note_row = 2 + len(rows) + 1
        ws.cell(row=note_row, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))
    return ws

# 1. Leads — verified live schema (PROJECT_AUDIT.md / CRM Adapter) — V1.1: AI-scoring columns removed
make_sheet(
    "Leads",
    ["Lead ID", "Date Created", "Source", "First Name", "Last Name", "Phone", "Email", "Address",
     "Service Needed", "Description", "Photos Link", "Preferred Time", "Status", "Last Contact",
     "Follow-up Count", "Assigned To", "Notes"],
    [
        ["EXAMPLE-LEAD-0001", "2026-06-01T09:14:00.000-04:00", "Website Form", "Maria", "Santos", "16175550111",
         "maria.santos@example.com", "42 Beacon St, Boston, MA", "Roof Inspection",
         "Noticed a leak near the chimney after last week's storm", "", "Weekday mornings",
         "Booked", "2026-06-01T09:20:00.000-04:00", 1, "J. Reyes",
         "Example row — replace with real data."],
        ["EXAMPLE-LEAD-0002", "2026-06-02T14:02:00.000-04:00", "Phone (Missed Call)", "David", "Chen", "16175550222",
         "", "", "", "", "", "", "New", "", 0, "",
         "Example row — missed-call auto-SMS sent (workflow 03); no Lead record is normally created for pure missed calls — this row illustrates what a follow-up form submission from the same number would look like once captured."],
    ],
    col_widths=[16, 18, 16, 12, 12, 14, 24, 24, 16, 32, 14, 16, 12, 18, 14, 12, 40],
    note=("SOURCE OF TRUTH: matches the live CRM Adapter sub-workflow contract exactly (workflow 01, "
          "id wVRHChyFrUNRaH4M) — verified against production data. 'Lead ID' and 'Phone' are the match "
          "keys; new leads get LEAD-#### minted by the adapter. Do not rename/remove columns — every "
          "workflow that follows up, reminds, or reports on leads depends on these exact header names. "
          "V1.1 (2026-06-11): 'Lead Score', 'Temperature', and 'Urgency' columns were removed system-wide "
          "— AI lead scoring was removed; every lead now follows the same 'Every Lead Alert' notification "
          "path (see docs/V1_1_RECONCILIATION.md). All timestamps are ISO 8601 in America/New_York "
          "(Boston time), the system-wide default timezone."),
)

# 2. Appointments — verified live schema (extracted from workflow 06 'Write Appointment' node)
make_sheet(
    "Appointments",
    ["Appt ID", "Lead ID", "Customer Name", "Phone", "Address", "Service Type", "Appt Date", "Appt Time",
     "Status", "Team Member", "Team Approval", "Calendar Event ID", "Reminder 24h", "Reminder 2h", "Notes",
     "Notified Appt Date", "Notified Appt Time"],
    [
        ["EXAMPLE-APT-20260603100000", "EXAMPLE-LEAD-0001", "Maria Santos", "16175550111",
         "42 Beacon St, Boston, MA", "Roof Inspection", "2026-06-10", "10:00 AM", "Scheduled",
         "J. Reyes", "", "", "", "", "Example row — booked via owner-facing form (workflow 06).",
         "2026-06-10", "10:00 AM"],
    ],
    col_widths=[24, 18, 16, 14, 24, 16, 12, 12, 12, 14, 14, 18, 14, 12, 40, 18, 18],
    note=("SOURCE OF TRUTH: matches the exact column set written by workflow 06's 'Write Appointment' "
          "node (live, verified) — 17 columns. 'Appt Date' must stay 'YYYY-MM-DD' and 'Appt Time' must "
          "stay 'H:MM AM/PM' — workflows 09 (Reminders), 11 (Health Monitor), and 13 (Reschedule "
          "Notifier) parse these with strict regexes. 'Reminder 24h'/'Reminder 2h' are write-only flag "
          "columns — leave them blank; the system manages them. 'Team Approval' and 'Calendar Event ID' "
          "are reserved for future phases (crew-approval workflow, calendar sync) — currently always "
          "blank. 'Notified Appt Date'/'Notified Appt Time' (added 2026-06-12) should be seeded equal "
          "to 'Appt Date'/'Appt Time' at booking time — workflow 13 compares against these to detect "
          "owner-initiated reschedules and notify the customer."),
)

# 3. Quotes — NOT YET BUILT. Reconstructed schema (no live workflow references this tab).
make_sheet(
    "Quotes",
    ["Quote ID", "Lead ID", "Customer Name", "Date Created", "Service Type", "Quote Amount", "Status",
     "Valid Until", "Sent Via", "Notes", "Created By"],
    [
        ["EXAMPLE-QUO-0001", "EXAMPLE-LEAD-0001", "Maria Santos", "2026-06-03", "Roof Replacement",
         8400.00, "Sent", "2026-07-03", "Email", "Full tear-off + architectural shingles, 30-yr warranty",
         "J. Reyes"],
    ],
    col_widths=[16, 18, 16, 14, 18, 14, 12, 14, 12, 40, 14],
    note=("⚠️ RECONSTRUCTED — NOT YET LIVE. No workflow currently reads or writes this tab. The original "
          "brief (Roofing_CRM_Google_Sheets.xlsx) named this tab but the file was never present in the "
          "project folder (see docs/CRM_SHEET_SCHEMA.md). This column set is a reasonable proposal for a "
          "future 'Quote Tracking' workflow — confirm with the client before building against it, and "
          "update this template + docs/CRM_SHEET_SCHEMA.md once a real schema is implemented and live."),
)

# 4. Jobs — NOT YET BUILT. Reconstructed schema.
make_sheet(
    "Jobs",
    ["Job ID", "Lead ID", "Appt ID", "Customer Name", "Phone", "Address", "Service Type", "Job Value",
     "Status", "Start Date", "Completion Date", "Assigned To", "Notes"],
    [
        ["EXAMPLE-JOB-0001", "EXAMPLE-LEAD-0001", "EXAMPLE-APT-20260603100000", "Maria Santos",
         "16175550111", "42 Beacon St, Boston, MA", "Roof Replacement", 8400.00, "Scheduled",
         "2026-06-17", "", "J. Reyes Crew", "Example row — converted from an accepted quote."],
    ],
    col_widths=[16, 18, 26, 16, 14, 24, 18, 12, 14, 14, 16, 16, 40],
    note=("⚠️ RECONSTRUCTED — NOT YET LIVE. No workflow currently reads or writes this tab. Referenced "
          "only as a future enhancement in docs/ROADMAP.md ('Job completion tracking in Jobs tab'). This "
          "column set anticipates a Phase-5-or-later 'Job Tracking' workflow that would convert booked "
          "appointments / accepted quotes into trackable jobs and feed completion data into the "
          "client-facing ROI report and case-study metrics (see docs/CASE_STUDY_DATA_PLAN.md, Metric 4 — "
          "'$ recovered' — which currently must be derived manually from Appointments + client-reported "
          "job value until this tab is wired up)."),
)

# 5. Communication Log — verified live schema (PROJECT_AUDIT.md / CRM Adapter)
make_sheet(
    "Communication Log",
    ["Log ID", "Date / Time", "Lead ID", "Customer Name", "Channel", "Direction", "Handler",
     "Message Summary", "Notes"],
    [
        ["EXAMPLE-LOG-20260601092000-481", "2026-06-01T09:20:00Z", "EXAMPLE-LEAD-0001", "Maria Santos",
         "SMS", "Outbound", "AI (Haiku 4.5)", "Sent confirmation + next-step SMS after form submission",
         "Example row — written by CRM Adapter on every inbound/outbound touch."],
        ["EXAMPLE-LOG-20260602140230-117", "2026-06-02T14:02:30Z", "", "David Chen", "Phone",
         "Inbound", "System", "Missed call — auto-SMS sent",
         "Example row — missed-call path; no Lead record created (skipLeadCreation routing)."],
    ],
    col_widths=[28, 20, 18, 16, 12, 12, 16, 44, 40],
    note=("SOURCE OF TRUTH: matches the live CRM Adapter contract exactly (workflow 01) — verified "
          "against production data. NOTE the header is 'Date / Time' WITH spaces around the slash — the "
          "adapter's column-mapping translates its internal 'Date/Time' (no spaces) key to this exact "
          "header. Do not 'fix' the spacing; it will break the column mapping."),
)

# 6. Follow Ups — NOT YET BUILT (workflow 05 tracks via 'Follow-up Count' on Leads, not this tab).
make_sheet(
    "Follow Ups",
    ["Follow-up ID", "Lead ID", "Customer Name", "Sequence Day", "Date Sent", "Channel",
     "Message Summary", "Response Received", "Notes"],
    [
        ["EXAMPLE-FU-0001", "EXAMPLE-LEAD-0002", "David Chen", "Day 1", "2026-06-03T09:00:00Z", "SMS",
         "First check-in: 'Still thinking about that roof inspection?'", "No",
         "Example row — illustrates a per-touch log entry; the live workflow currently records only an "
         "aggregate counter on the Leads tab (see note)."],
    ],
    col_widths=[16, 18, 16, 14, 20, 12, 44, 16, 44],
    note=("⚠️ RECONSTRUCTED — NOT YET LIVE AS A SEPARATE TAB. The live Follow-Up Sequence workflow "
          "(05, chYfABnQdnPfiHQx) currently tracks progress via a single 'Follow-up Count' integer "
          "column on the Leads tab plus per-touch entries in Communication Log — it does NOT write to "
          "a dedicated 'Follow Ups' tab. This reconstructed schema proposes what a per-touch audit log "
          "would look like if a future enhancement (e.g. more granular reporting, A/B testing different "
          "sequences) needed one. Until then, treat 'Follow-up Count' (Leads) + Communication Log as "
          "the system of record for follow-up activity."),
)

# 7. Team Schedule — NOT YET BUILT. Reconstructed schema.
make_sheet(
    "Team Schedule",
    ["Date", "Team Member", "Shift / Availability", "Assigned Jobs", "Notes"],
    [
        ["2026-06-17", "J. Reyes Crew", "8:00 AM – 4:00 PM", "EXAMPLE-JOB-0001",
         "Example row — illustrates a daily crew-availability + assignment view."],
    ],
    col_widths=[14, 18, 20, 28, 44],
    note=("⚠️ RECONSTRUCTED — NOT YET LIVE. No workflow currently reads or writes this tab. Named in "
          "the original brief as part of the 8-tab CRM spec, but never built against — there is no "
          "live crew-scheduling or calendar-sync workflow yet (see docs/ROADMAP.md 'Calendar sync' "
          "under optional enhancements / Phase 5+). This column set is a starting proposal only; "
          "confirm the client's actual scheduling workflow before building anything against it — "
          "field-service businesses vary widely here (some use paper boards, some use dedicated apps "
          "like Housecall Pro / ServiceTitan, which may make this tab redundant for some clients)."),
)

# 8. Dashboard — formula/summary tab, not a data-entry tab.
ws = make_sheet(
    "Dashboard",
    ["Metric", "Value", "Notes"],
    [
        ["Total Active Leads", 0,
         "PLACEHOLDER — replace with e.g. =COUNTIFS(Leads!M:M,\"<>Booked\",Leads!M:M,\"<>Lost\") once real data exists. ('Status' is column M in the V1.1 17-column Leads schema.)"],
        ["New Leads This Week", 0, "PLACEHOLDER — e.g. =COUNTIFS(Leads!B:B, \">=\"&TODAY()-7)"],
        ["Appointments This Week", 0, "PLACEHOLDER — e.g. =COUNTIFS(Appointments!G:G, \">=\"&TEXT(TODAY()-7,\"YYYY-MM-DD\"))"],
        ["Jobs Completed This Month", 0,
         "PLACEHOLDER — depends on the Jobs tab going live first (see that tab's note); e.g. =COUNTIFS(Jobs!I:I,\"Completed\",Jobs!K:K,\">=\"&EOMONTH(TODAY(),-1)+1)"],
        ["Stale Leads (No Contact 7+ Days)", 0,
         "PLACEHOLDER — mirrors workflow 07's own 'Stale' definition; e.g. =COUNTIFS(Leads!N:N,\"<\"&TEXT(TODAY()-7,\"YYYY-MM-DD\"))  — keep this formula in sync with whatever workflow 07 (Pipeline Status Digest) uses internally so the dashboard never disagrees with its email digest. ('Last Contact' is column N in the V1.1 17-column Leads schema.)"],
    ],
    col_widths=[34, 12, 90],
    note=("⚠️ THIS TAB IS A SUMMARY VIEW, NOT A DATA-ENTRY TAB — it is meant to hold live formulas that "
          "reference the other tabs, not raw rows. No workflow currently writes to or reads from it; "
          "the system's owner-facing reporting is delivered instead via email digests (workflows 07/08, "
          "converted from SMS to email in V1.1) and Workflow 11's health alerts. The 'Value' column above "
          "is intentionally left at 0/placeholder — replace each with a live formula (see the Notes "
          "column for a starting point per metric) once the sheet has real rows to compute against. Keep "
          "every formula's definition of a status ('Stale', etc.) in sync with the corresponding live "
          "workflow's own definition — a Dashboard that disagrees with the email digest the owner already "
          "trusts is worse than no Dashboard at all. V1.1 (2026-06-11): the 'Hot/Emergency Leads (Open)' "
          "metric was removed along with AI lead scoring system-wide."),
)
for row in ws.iter_rows(min_row=2, max_row=7, min_col=2, max_col=2):
    for cell in row:
        cell.number_format = "#,##0"

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
